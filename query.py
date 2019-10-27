import os
import django
import openpyxl
from openpyxl.worksheet import worksheet
import sys

reload(sys)
sys.setdefaultencoding('utf8')

os.environ["PYTHONIOENCODING"] = "utf-8"
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sracademic.settings")
django.setup()
from django.contrib.auth.models import User
from django.db import transaction
from sracademic.settings import BASE_DIR
from academic.models import *
from datetime import date


def del_espace(element):
    element = element.strip()
    newelement = ''
    lenght = len(element) - 1
    count = 0
    while count <= lenght:
        if element[count] == ' ':
            if element[count + 1] != ' ':
                newelement += element[count]
        else:
            newelement += element[count]
        count += 1
    return newelement


currentValues = []
line = 1
currentValues = []
fechaIngreso = date(2019, 6, 18)
try:
    archivo = BASE_DIR + '\\Datos.xlsx'
    wb = openpyxl.load_workbook(archivo)
    worksheet = wb["MATERIAS"]
    with transaction.atomic():
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            materias = Materia(nombre=currentValues[0], alias=currentValues[1], credito=currentValues[2])
            materias.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["PERSONA"]
        fechanacimiento = date(2019, 6, 18)
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Persona(nombres=currentValues[0], apellidos=currentValues[1],
                              documento=currentValues[2], correo=currentValues[3],
                              fechaNacimiento=fechanacimiento,
                              celular=currentValues[5], telefono=currentValues[6],
                              direccion=currentValues[7])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["ESTUDIANTE"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Estudiante(persona_id=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["DOCENTE"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Docente(persona_id=currentValues[0], foto="")
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["FACULTAD"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Facultad(nombre=currentValues[0], alias=currentValues[1])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["PARALELO"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Paralelo(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["SEMESTRE"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Semestre(nombre=currentValues[0], alias=currentValues[1])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["CARRERA"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Carrera(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["SESION"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Sesion(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["AREAC"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = AreaConocimiento(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["GRADO"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Grado(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

            worksheet = wb["TIPONIVEL"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = TipoNivel(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["MODALIDAD"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Modalidad(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["SUBAREACE"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = SubAreaConocimientoEspecifico(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["SUBAREAC"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = SubAreaConocimiento(nombre=currentValues[0])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["TITULACION"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Titulacion(docente_id=currentValues[0], descripcion=currentValues[1],
                                 tiponivel_id=currentValues[2], grado_id=currentValues[3],
                                 areaconocimiento_id=currentValues[4], subareaconocimiento_id=currentValues[5],
                                 areaconocimientoespecifico_id=currentValues[6])
            persona.save()

        worksheet = wb["NIVELACADEMICO"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = NivelAcademico(nombre=currentValues[0], modalidad_id=currentValues[1],
                                     matriculas=currentValues[2], sesion_id=currentValues[3],
                                     fechainicio=date(2019, 5, 4), fechafin=date(2019, 9, 27),
                                     fechamatriculaordinaria=date(2019, 9, 6),
                                     fechamatriculaextraordinaria=date(2019, 9, 10),
                                     fechamatriculaespecial=date(2019, 10, 18))
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["PERIODOACADEMICO"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = PeriodoAcademico(descripcion=currentValues[0], fechaInicio=date(2019, 9, 10),
                                       fechaFin=date(2019, 10, 18))
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["EVALUACION"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Evaluacion(estudiante_id=currentValues[0], docente_id=currentValues[1],
                                 materia_id=currentValues[2], puntuacion=currentValues[3])
            persona.save()
            print(str(currentValues))

            counter += 1
            line += 1

        worksheet = wb["INVESTIGACION"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Investigacion(docente_id=currentValues[0], descripcion=currentValues[1],
                                    areaconocimiento_id=currentValues[2],
                                    subareaconocimiento_id=currentValues[3],
                                    areaconocimientoespecifico_id=currentValues[4])
            persona.save()

        worksheet = wb["DISTRIBUTIVO"]
        for row in worksheet.iter_rows(min_row=2):
            currentValues = []
            counter = 0
            for cell in row:
                currentValues.append(str(cell.value))
            persona = Distributivo(peridoacademico_id=currentValues[0],
                                   facultad_id=currentValues[1],
                                   carera_id=currentValues[2],
                                   semestre_id=currentValues[3],
                                   materia_id=currentValues[4],
                                   paralelo_id=currentValues[5],
                                   docente_id=currentValues[6],
                                   fechainicio=date(2019, 5, 30),
                                   fechafin=date(2019, 9, 27),
                                   cupo=currentValues[9],
                                   horassemanal=currentValues[10],
                                   sesion_id=currentValues[11])
            persona.save()






except Exception as ex:
    print(ex)
    print(str(line))
    print(currentValues)
